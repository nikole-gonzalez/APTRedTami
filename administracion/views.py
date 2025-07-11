
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import auth, messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.db.models import Q

import base64
from datetime import datetime, date
import matplotlib.pyplot as plt
import matplotlib
matplotlib.use('Agg')
from io import BytesIO
import requests
import numpy as np
from django.contrib.sessions.backends.db import SessionStore
from django.contrib import auth, messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.db import connection
from django.db.models import Count, F, Max
from django.http import HttpResponse, JsonResponse, HttpResponseForbidden
from django.utils import timezone

from collections import Counter
from django.db.models.functions import ExtractYear
from django.db.models.functions import TruncDate


from openpyxl import Workbook

from .models import *
from .forms import *

from usuario.models import Agenda, Cesfam
from administracion.models import Usuario

from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from reportlab.lib.pagesizes import letter, landscape, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle


from django.views.decorators.csrf import csrf_exempt
import json
import locale 

import logging

from django.conf import settings
from django.db.models import F, Subquery, OuterRef, Case, When, Value, CharField
from datetime import date

from .utils import *
from django.db.models.functions import Concat
from django.views.decorators.cache import never_cache


locale.setlocale(locale.LC_TIME, 'es_ES')

@login_required(login_url='/login/')
def home(request):
    return render(request, 'administracion/index.html')

@login_required(login_url='/login/')
def admin_index(request):
    return render(request, 'administracion/index.html')

def cesfam_index(request):
    return render(request, 'cesfam/index_c.html')

@login_required
def respuestas(request):
    return render(request, 'administracion/respuestas.html')

@login_required
def opc_vis_agenda(request):
    return render(request, 'administracion/opc_vis_agenda.html')

@login_required
def historial_agendamientos(request):
    return render(request, 'administracion/historial_agendamientos.html')

@login_required
def gestion_usuarios(request):
    return render(request, 'administracion/gestion_usuarios.html')

@login_required
def opc_vis_FRM(request):
    return render(request, 'administracion/opc_vis_FRM.html')

@login_required
def opc_vis_FRNM(request):
    return render(request, 'administracion/opc_vis_FRNM.html')

@login_required
def opc_vis_DS(request):
    return render(request, 'administracion/opc_vis_DS.html')

# ------------------------------------------------------ #
# ---------------------- Reportes ---------------------- #
# ------------------------------------------------------ #

@login_required
def reportes(request):
    grafico_genero = generar_grafico_personas_por_genero()
    grafico_comuna = generar_grafico_ingresos_por_comuna()
    grafico_pap_tres_anios = generar_grafico_realizado_pap_tres_anios()
    grafico_escolaridad = generar_grafico_escolaridad()
    grafico_anio_nac = generar_grafico_anio_nacimiento()
    grafico_resp_diarias = generar_grafico_respuestas_por_dia()
    grafico_usuarias_edad = generar_grafico_usuario_por_edad()
    grafico_usuarias_cesfam = generar_grafico_usuarios_por_cesfam()
    grafico_ingresos_diarios_cesfam = generar_graficos_ingresos_diarios_por_cesfam()
    grafico_realizado_pap_cesfam = generar_grafico_realizado_pap_por_cesfam()
    data = {
        "imagen_base64_personas_por_genero": grafico_genero,
        "imagen_base64_ingresos_por_comuna": grafico_comuna,
        "imagen_base64_realizado_pap_tres_anios": grafico_pap_tres_anios,
        "imagen_base64_escolaridad": grafico_escolaridad,
        "imagen_base64_anio_nacimiento": grafico_anio_nac,
        "imagen_base64_resp_por_dia": grafico_resp_diarias,
        "imagen_base64_usuarias_por_edad": grafico_usuarias_edad,
        "imagen_base64_usuarias_por_cesfam": grafico_usuarias_cesfam,
        "imagen_base64_ingresos_diarios_por_cesfam": grafico_ingresos_diarios_cesfam,
        "imagen_base64_pap_por_cesfam": grafico_realizado_pap_cesfam,
        "hay_datos": grafico_genero or grafico_comuna or grafico_pap_tres_anios or grafico_escolaridad
            or grafico_anio_nac or grafico_resp_diarias or grafico_usuarias_edad or grafico_usuarias_cesfam or 
            grafico_ingresos_diarios_cesfam
    }
    return render(request, 'administracion/reportes.html', data)

def convertir_grafico_a_base64():
    buffer = BytesIO()
    plt.savefig(buffer, format="png", dpi=100, bbox_inches='tight')
    plt.close()
    buffer.seek(0)
    return base64.b64encode(buffer.getvalue()).decode("utf-8")

def calcular_edad(fecha_nacimiento):
    today = date.today()
    return today.year - fecha_nacimiento.year - ((today.month, today.day) < (fecha_nacimiento.month, fecha_nacimiento.day))

plt.rcParams['font.family'] = 'sans-serif'  
plt.rcParams['font.sans-serif'] = 'Calibri' 
plt.rcParams['font.size'] = 14
plt.rcParams['axes.titlesize'] =23
plt.rcParams['axes.labelsize']= 14
plt.rcParams['axes.labelpad']=10

def generar_grafico_personas_por_genero():
    respuestas = RespFRNM.objects.filter(id_opc_frnm__id_opc_frnm__in=[1, 2, 3])

    if not respuestas.exists():
        return None

    contador = Counter(respuestas.values_list('id_opc_frnm__id_opc_frnm', flat=True))

    generos = []
    cantidades = []

    for genero_id in [1, 2, 3]:
        try:
            genero = OpcFRNM.objects.get(id_opc_frnm=genero_id)
            generos.append(genero.opc_resp_frnm)
            cantidades.append(contador.get(genero_id, 0))
        except OpcFRNM.DoesNotExist:
            continue

    if not any(cantidades):
        return None

    colores = {'Masculino': '#79addc', 'Femenino': '#EFB0C9', 'Otro': '#A5F8CE'}

    plt.figure(figsize=(8, 6))

    plt.bar(generos, cantidades, color=[colores.get(g, '#CCCCCC') for g in generos])

    for i in range(len(generos)):
        plt.text(i, cantidades[i], str(cantidades[i]), ha='center', va='bottom')

    plt.xlabel("Género")
    plt.ylabel("Número de Personas")
    plt.ylim(top=max(cantidades) + 2) 
    plt.title("Ingresos por Género", pad=20)

    return convertir_grafico_a_base64()

def generar_grafico_ingresos_por_comuna():
    
    respuestas = RespFRNM.objects.filter(id_opc_frnm__id_opc_frnm__in=[1, 4])

    datos_agrupados = (
        respuestas
        .values('id_manychat__cod_comuna__nombre_comuna')
        .annotate(total=Count('id_resp_frnm'))
        .order_by('id_manychat__cod_comuna__nombre_comuna')
    )

    comunas = [dato['id_manychat__cod_comuna__nombre_comuna'] for dato in datos_agrupados]
    total_ingresos = [dato['total'] for dato in datos_agrupados]

    if not total_ingresos:
        return None  

    # Crear gráfico circular
    fig, ax = plt.subplots()
    wedges, texts, autotexts = ax.pie(
        total_ingresos,
        labels=comunas,
        autopct=lambda pct: f"{pct:.1f}%\n{int(pct/100 * sum(total_ingresos) + 0.5)} ingresos",
        startangle=90
    )
    ax.axis('equal')
    ax.set_title('Distribución de Ingresos por Comuna', pad=20)

    # Ajustar el tamaño del texto
    for text, autotext in zip(texts, autotexts):
        text.set(size=8)
        autotext.set(size=8)

    return convertir_grafico_a_base64()


def generar_grafico_realizado_pap_tres_anios():
    
    manychat_ids = RespFRNM.objects.filter(
        id_opc_frnm__id_opc_frnm__in=[1, 4]
    ).values_list('id_manychat', flat=True)

   
    respuestas_tm = RespTM.objects.filter(
        id_manychat__in=manychat_ids,
        id_opc_tm__id_opc_tm__in=[1, 2, 3, 4]
    )

    contador = Counter(respuestas_tm.values_list('id_opc_tm__id_opc_tm', flat=True))

    if not contador:
        return None

    labels = []
    sizes = []
    counts = []

    for id_opc_tm in [1, 2, 3,4]:
        try:
            opcion = OpcTM.objects.get(id_opc_tm=id_opc_tm)
            cantidad = contador.get(id_opc_tm, 0)
            labels.append(opcion.opc_resp_tm)
            sizes.append(cantidad)
            counts.append(f"{opcion.opc_resp_tm} - {cantidad}")
        except OpcTM.DoesNotExist:
            continue

    # Crear gráfico circular
    fig, ax = plt.subplots(figsize=(8, 8))
    wedges, texts, autotexts = ax.pie(
        sizes, labels=None, autopct='%1.1f%%', startangle=90,
        colors=['#79addc', '#EFB0C9', '#A5F8CE', '#FFD166']
    )
    
    ax.legend(wedges, counts, title="Respuestas", loc="center left", bbox_to_anchor=(1, 0, 0.5, 1))
    plt.title('¿Te has realizado una PAP en los últimos 3 años?', pad=20)

    return convertir_grafico_a_base64()

def generar_grafico_escolaridad():
   
    manychat_ids = RespFRNM.objects.filter(
        id_opc_frnm__id_opc_frnm__in=[1, 4]
    ).values_list('id_manychat', flat=True)

    respuestas_ds = RespDS.objects.filter(
        id_manychat__in=manychat_ids,
        id_opc_ds__id_opc_ds__in=[1, 2, 3]
    )

    contador = Counter(respuestas_ds.values_list('id_opc_ds__id_opc_ds', flat=True))

    if not contador:
        return None

    labels = []
    sizes = []
    counts = []

    for id_opc_ds in [1, 2, 3]:
        try:
            opcion = OpcDS.objects.get(id_opc_ds=id_opc_ds)
            cantidad = contador.get(id_opc_ds, 0)
            labels.append(opcion.opc_resp_ds)
            sizes.append(cantidad)
            counts.append(f"{opcion.opc_resp_ds} - {cantidad}")
        except OpcDS.DoesNotExist:
            continue

    fig, ax = plt.subplots(figsize=(8, 8))
    wedges, texts, autotexts = ax.pie(
        sizes,
        labels=None,
        autopct='%1.1f%%',
        startangle=90,
        colors=['#79addc', '#EFB0C9', '#A5F8CE']
    )

    ax.legend(wedges, counts, title="Respuestas", loc="center left", bbox_to_anchor=(1, 0, 0.5, 1))
    plt.title('Nivel de estudio usuarias', pad=20)

    return convertir_grafico_a_base64()

def generar_grafico_anio_nacimiento():
    
    manychat_ids = RespFRNM.objects.filter(
        id_opc_frnm__id_opc_frnm__in=[1, 4]
    ).values_list('id_manychat', flat=True)

    datos = (
        Usuario.objects
        .filter(id_manychat__in=manychat_ids)
        .annotate(anio=ExtractYear('fecha_nacimiento'))
        .values('anio')
        .annotate(cantidad=Count('id_manychat'))
        .order_by('anio')
    )

    anios = [dato['anio'] for dato in datos]
    cantidades = [dato['cantidad'] for dato in datos]

    if not anios:
        return None

    plt.figure(figsize=[18, 8])
    plt.bar(anios, cantidades, color="#79addc")
    plt.xlabel("Año de Nacimiento")
    plt.ylabel("Número de Usuarios")
    plt.yticks(range(0, max(cantidades) + 1))
    plt.ylim(top=max(cantidades) + 2) 
    plt.title("Usuarias por Año de Nacimiento", pad=20)
    plt.xticks(range(min(anios), max(anios)+1, 1), rotation=90)

    for anio, cantidad in zip(anios, cantidades):
        plt.text(anio, cantidad, str(cantidad), ha='center', va='bottom')

    return convertir_grafico_a_base64()

def generar_grafico_respuestas_por_dia():
   
    manychat_ids = RespFRNM.objects.filter(
        id_opc_frnm__id_opc_frnm__in=[1, 4]
    ).values_list('id_manychat', flat=True)

    datos = (
        Usuario.objects
        .filter(id_manychat__in=manychat_ids)
        .annotate(fecha=TruncDate('fecha_ingreso'))
        .values('fecha')
        .annotate(cantidad=Count('id_manychat'))
        .order_by('fecha')
    )

    fechas = [dato['fecha'].strftime("%d-%m-%Y") for dato in datos]
    cantidades = [dato['cantidad'] for dato in datos]

    if not fechas:
        return None

    plt.figure(figsize=[14, 6])
    plt.plot(fechas, cantidades, marker="o", linestyle="-", color="#79addc")
    plt.xlabel("Fecha de Respuesta")
    plt.ylabel("Número de Respuestas")
    plt.yticks(range(0, max(cantidades) + 1))
    plt.ylim(top=max(cantidades) + 2) 
    plt.title("Respuestas por Día", pad=20)
    plt.xticks(rotation=90)
    plt.tight_layout()

    for fecha, cantidad in zip(fechas, cantidades):
        plt.annotate(f"{cantidad}", (fecha, cantidad), textcoords="offset points", xytext=(0,10), ha='center')

    return convertir_grafico_a_base64()

def generar_grafico_usuario_por_edad():
    usuarios = Usuario.objects.filter(
        respfrnm__id_opc_frnm__in=[1,4]
    ).values_list('fecha_nacimiento', flat=True)

    # Calcular edades
    edades = [calcular_edad(fn) for fn in usuarios if fn]
    
    if not edades:
        return None

    # Contar ocurrencias por edad
    contador = Counter(edades)
    edades_ordenadas = sorted(contador.keys())
    cantidades = [contador[edad] for edad in edades_ordenadas]

    # Crear gráfico
    plt.figure(figsize=[18, 8])
    plt.bar(edades_ordenadas, cantidades, color="#79addc")
    plt.xlabel("Edad")
    plt.ylabel("Número de Usuarias")
    plt.yticks(range(0, max(cantidades) + 1))
    plt.ylim(top=max(cantidades) + 2) 
    plt.title("Usuarias por edad", pad=20)
    
    if edades_ordenadas:
        plt.xticks(range(min(edades_ordenadas), max(edades_ordenadas) + 1, 1))

    for edad, cantidad in zip(edades_ordenadas, cantidades):
        plt.text(edad, cantidad, str(cantidad), ha='center', va='bottom')

    return convertir_grafico_a_base64()

def generar_grafico_usuarios_por_cesfam():
    datos = (
        Usuario.objects
        .filter(cesfam_usuario__isnull=False)  
        .values('cesfam_usuario__nombre_cesfam')  
        .annotate(total=Count('id_manychat'))  
        .order_by('-total')  
    )

    cesfams = [dato['cesfam_usuario__nombre_cesfam'] for dato in datos]
    totales = [dato['total'] for dato in datos]

    if not cesfams:
        return None  

    plt.figure(figsize=(12, 7))

    colores = ['#79addc', '#EFB0C9', '#A5F8CE', '#FFD166', '#06D6A0']
    
    barras = plt.bar(cesfams, totales, color=colores[:len(cesfams)])
    
    for barra in barras:
        altura = barra.get_height()
        plt.text(barra.get_x() + barra.get_width()/2., altura,
                f'{int(altura)}',
                ha='center', va='bottom')
    
    plt.xlabel('CESFAM', labelpad=15)
    plt.ylabel('Número de Usuarios', labelpad=15)
    plt.yticks(range(0, max(totales) + 1))
    plt.ylim(top=max(totales) + 2) 
    plt.title('Distribución de Usuarios por CESFAM', pad=20)
    plt.xticks(rotation=45, ha='right') 
    
    plt.tight_layout()
    
    return convertir_grafico_a_base64()

def generar_graficos_ingresos_diarios_por_cesfam():
    try:
        cesfams = Cesfam.objects.all()
        
        if not cesfams:
            print("No hay CESFAMs registrados en el sistema")
            return None

        graficos_por_cesfam = {}
        colores = ['#4e79a7', '#e15759', '#76b7b2', '#59a14f', '#edc948']
        
        for i, cesfam in enumerate(cesfams):
            try:
                datos = (
                    Usuario.objects
                    .filter(cesfam_usuario__id_cesfam=cesfam.id_cesfam, fecha_ingreso__isnull=False)
                    .extra({'fecha': "DATE(fecha_ingreso)"})   
                    .values('fecha')
                    .annotate(cantidad=Count('id_manychat'))
                    .order_by('fecha')
                )

                datos_validos = [d for d in datos if d['fecha'] is not None]
                
                if not datos_validos:
                    print(f"No hay datos válidos para {cesfam.nombre_cesfam}")
                    plt.figure(figsize=(12, 6))
                    plt.text(0.5, 0.5, 'Sin datos disponibles', 
                            ha='center', va='center', fontsize=12)
                    plt.title(f"Ingresos Diarios - {cesfam.nombre_cesfam}", pad=20)
                    graficos_por_cesfam[cesfam.nombre_cesfam] = convertir_grafico_a_base64()
                    continue

                fechas = [d['fecha'].strftime("%d-%m-%Y") for d in datos_validos]
                cantidades = [d['cantidad'] for d in datos_validos]

                plt.figure(figsize=(12, 6))
                plt.plot(fechas, cantidades, marker="o", linestyle="-", 
                        color=colores[i % len(colores)], linewidth=2.5, markersize=8)
                
                plt.xlabel("Fecha", fontsize=12)
                plt.ylabel("Ingresos Diarios", fontsize=12)
                plt.yticks(range(0, max(cantidades) + 1)) 
                plt.ylim(top=max(cantidades) + 2) 
                plt.title(f"Ingresos Diarios - {cesfam.nombre_cesfam}", pad=20, fontsize=14)
                plt.xticks(rotation=45, ha='right')
                plt.grid(True, linestyle='--', alpha=0.3)
                
                for fecha, cantidad in zip(fechas, cantidades):
                    plt.annotate(
                        f"{cantidad}",
                        (fecha, cantidad),
                        textcoords="offset points",
                        xytext=(0, 10),
                        ha='center',
                        fontsize=9
                    )
                
                plt.tight_layout()
                graficos_por_cesfam[cesfam.nombre_cesfam] = convertir_grafico_a_base64()

            except Exception as e:
                print(f"Error generando gráfico para {cesfam.nombre_cesfam}: {str(e)}")
                continue

        return graficos_por_cesfam if graficos_por_cesfam else None

    except Exception as e:
        print(f"Error general en generar_grafico_ingresos_diarios_por_cesfam: {str(e)}")
        return None

def generar_grafico_realizado_pap_por_cesfam():
    try:
        cesfams = Cesfam.objects.all()
        if not cesfams:
            print("No hay CESFAMs registrados.")
            return None

        etiquetas = []
        cantidades = []

        for cesfam in cesfams:
            usuarios_ids = Usuario.objects.filter(
                cesfam_usuario=cesfam
            ).values_list('id_manychat', flat=True)

            cantidad_si = RespTM.objects.filter(
                id_manychat__in=usuarios_ids,
                id_opc_tm__id_opc_tm=1
            ).count()

            if cantidad_si > 0:
                etiquetas.append(cesfam.nombre_cesfam)
                cantidades.append(cantidad_si)

        if not cantidades:
            print("No hay datos para mostrar el gráfico de PAP.")
            return None

        fig, ax = plt.subplots(figsize=(8, 8))
        wedges, texts, autotexts = ax.pie(
            cantidades,
            labels=None,
            autopct='%1.1f%%',
            startangle=90,
            colors=['#79addc', '#EFB0C9', '#A5F8CE', '#FFD166', '#06D6A0']
        )

        leyenda = [f"{nombre} ({cantidad})" for nombre, cantidad in zip(etiquetas, cantidades)]

        ax.legend(wedges, leyenda, title="CESFAM", loc="center left", bbox_to_anchor=(1, 0, 0.5, 1))
        plt.title('Usuarios que se realizaron PAP en los últimos 3 años por CESFAM', pad=20)
        plt.tight_layout()

        return convertir_grafico_a_base64()

    except Exception as e:
        print(f"Error al generar gráfico de PAP por CESFAM: {str(e)}")
        return None
    
# ------------------------------------------------------------------ #
# ---------------------- Respuestas Usuarias ----------------------- #
# ------------------------------------------------------------------ #

def ajustar_ancho_columnas(ws):
  
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)

        for cell in col:
            try:
                if cell.value: 
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        
        ws.column_dimensions[col_letter].width = max_length + 2

def background_colors(ws):

    color = "00a0b3a8" 
    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    for cell in ws[1]: 
        cell.fill = fill


def datos_perfil(request):
    query = request.GET.get("q", "").strip()
    
    datos_query = Usuario.objects.annotate(
        rut_completo=Concat(
            F('rut_usuario'),
            Value('-'),
            F('dv_rut'),
            output_field=CharField()
        ),
        rut_sin_formato=Concat(
            F('rut_usuario'),
            F('dv_rut'),
            output_field=CharField()
        )
    ).order_by("-fecha_ingreso")

    if query:
        cleaned_query = query.replace(".", "").replace("-", "").replace(" ", "").lower()
        
        filtros = Q(rut_completo__icontains=query) | Q(rut_sin_formato__icontains=cleaned_query)

        if len(cleaned_query) > 1:
            filtros |= Q(rut_usuario__icontains=cleaned_query[:-1])
        
        filtros |= Q(id_manychat__icontains=query) | Q(num_whatsapp__icontains=query)
        
        datos_query = datos_query.filter(filtros)
    
    page_obj = paginacion_queryset1(request, datos_query)

    return render(request, 'administracion/datos_perfil.html', {
        "page_obj": page_obj,
        "query": query,
    })
# ------------------ #
# ---- Tamizaje ---- #
# ------------------ #

@login_required
def tamizaje(request):
    query = request.GET.get("q", "").strip()

    datos_query = RespTM.objects.select_related(
        "id_opc_tm", "id_opc_tm__id_preg_tm", "id_manychat"
    ).order_by("-fecha_respuesta_tm")


    if query:
        datos_query = filtrar_por_rut_o_manychat(datos_query, query)

    datos_query = datos_query.values(
        "id_resp_tm",
        "id_manychat",
        "id_opc_tm__id_preg_tm__preg_tm",
        "id_opc_tm__opc_resp_tm",
        "fecha_respuesta_tm",
        "id_manychat__rut_usuario",
        "id_manychat__dv_rut"
    )

    page_obj = paginacion_queryset1(request, datos_query)

    return render(request, 'administracion/tamizaje.html', {
        "page_obj": page_obj,
        "query": query
    })

#Validar contraseña de descargas
def validar_password(request):
    if request.method == 'POST':
        password = request.POST.get('password')
        if password == settings.DOWNLOAD_PASSWORD:
            session = SessionStore()
            session['password_validated'] = True
            session.set_expiry(300) 
            session.save()
            return JsonResponse({
                'valid': True, 
                'session_key': session.session_key
            })
        else:
            return JsonResponse({'valid': False, 'error': 'Contraseña incorrecta'})
    return JsonResponse({'valid': False, 'error': 'Método no permitido'})

def crear_excel_datos_tamizaje(request):
    session_key = request.GET.get('session_key')
    if session_key:
        session = SessionStore(session_key=session_key)
        if session.get('password_validated', False):
            session.flush()

            wb = Workbook()
            ws = wb.active
            ws.title = "Resultados Tamizaje"

            ws.append(['Rut', 'Pregunta', 'Respuesta', 'Fecha Respuesta'])

            respuestas = RespTM.objects.select_related(
                'id_opc_tm__id_preg_tm', 'id_manychat'
            ).values(
                'id_manychat__rut_usuario',
                'id_manychat__dv_rut',
                'id_opc_tm__id_preg_tm__preg_tm',
                'id_opc_tm__opc_resp_tm',
                'fecha_respuesta_tm'
            ).order_by('-fecha_respuesta_tm')

            for r in respuestas:
                fecha = r['fecha_respuesta_tm']
                fecha_str = fecha.strftime('%d-%m-%Y %H:%M:%S') if fecha else ''
                
                ws.append([
                    f"{r['id_manychat__rut_usuario']}-{r['id_manychat__dv_rut']}",
                    r['id_opc_tm__id_preg_tm__preg_tm'],
                    r['id_opc_tm__opc_resp_tm'],
                    fecha_str,
                ])

            ajustar_ancho_columnas(ws)
            background_colors(ws)

            response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            response['Content-Disposition'] = 'attachment; filename="datos_tamizaje.xlsx"'
            
            wb.save(response)
            return response
    return HttpResponseForbidden("Acceso no autorizado")

def crear_pdf_datos_tamizaje(request):
    session_key = request.GET.get('session_key')
    if session_key:
        session = SessionStore(session_key=session_key)
        if session.get('password_validated', False):
            session.flush()

            def truncate_text(text, max_length):
                if not text:
                    return text
                return (text[:max_length-3] + '...') if len(text) > max_length else text

            buffer = BytesIO()
            doc = SimpleDocTemplate(
                buffer,
                pagesize=landscape(A4),
                leftMargin=1*cm,
                rightMargin=1*cm,
                topMargin=1.5*cm,
                bottomMargin=1.5*cm
            )
            
            styles = getSampleStyleSheet()
            styles.add(ParagraphStyle(
                name='Small',
                parent=styles['Normal'],
                fontSize=7,
                leading=9
            ))

            respuestas = RespTM.objects.select_related(
                'id_opc_tm__id_preg_tm', 'id_manychat'
            ).values(
                'id_manychat__rut_usuario',
                'id_manychat__dv_rut',
                'id_opc_tm__id_preg_tm__preg_tm',
                'id_opc_tm__opc_resp_tm',
                'fecha_respuesta_tm'
            ).order_by('id_manychat__rut_usuario')

            # Procesamiento de datos
            dict_respuestas = {}
            for respuesta in respuestas:
                rut = f"{respuesta['id_manychat__rut_usuario']}-{respuesta['id_manychat__dv_rut']}"
                pregunta = respuesta['id_opc_tm__id_preg_tm__preg_tm']
                respuesta_usuario = respuesta['id_opc_tm__opc_resp_tm']
                fecha = respuesta['fecha_respuesta_tm']
                
                if rut not in dict_respuestas:
                    dict_respuestas[rut] = {
                        'fecha': fecha.strftime('%d-%m-%Y %H:%M:%S') if fecha else 'Sin fecha', 
                        'respuestas': {}
                    }
                dict_respuestas[rut]['respuestas'][pregunta] = respuesta_usuario

            encabezados = ['RUT', 'Pregunta', 'Respuesta', 'Fecha Respuesta']
            data = [encabezados]

            for rut, datos in dict_respuestas.items():
                for pregunta, respuesta in datos['respuestas'].items():
                    data.append([
                        rut,
                        truncate_text(pregunta, 40),
                        truncate_text(respuesta, 30),
                        datos['fecha']
                    ])

            tabla = Table(data, repeatRows=1)
            
            # Calcular ancho de columnas
            ancho_total = landscape(A4)[0] - 2*cm 
            ancho_rut = 4*cm
            ancho_fecha = 3*cm
            ancho_pregunta = (ancho_total - ancho_rut - ancho_fecha) * 0.6
            ancho_respuesta = (ancho_total - ancho_rut - ancho_fecha) * 0.4
            
            estilo = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#c6fffa')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('FONTSIZE', (0, 1), (-1, -1), 7),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 0.25, colors.lightgrey),
                ('WORDWRAP', (0, 0), (-1, -1), True),
            ])
            
            # Aplicar anchos de columnas
            estilo.add('COLWIDTH', (0, 0), (0, -1), ancho_rut)
            estilo.add('COLWIDTH', (1, 0), (1, -1), ancho_pregunta)
            estilo.add('COLWIDTH', (2, 0), (2, -1), ancho_respuesta)
            estilo.add('COLWIDTH', (3, 0), (3, -1), ancho_fecha)
            
            # Filas alternadas
            for i in range(1, len(data)):
                if i % 2 == 0:
                    estilo.add('BACKGROUND', (0, i), (-1, i), colors.whitesmoke)
            
            tabla.setStyle(estilo)

            elementos = [
                Paragraph("Tamizaje", styles['Title']),
                Spacer(1, 0.5*cm),
                Paragraph(f"Total de respuestas: {len(data)-1}", styles['Normal']),
                Spacer(1, 0.5*cm),
                tabla,
                Spacer(1, 0.3*cm),
                Paragraph(f"Generado el: {timezone.now().strftime('%d-%m-%Y %H:%M:%S')}", styles['Small'])
            ]

            doc.build(elementos)
            
            buffer.seek(0)
            response = HttpResponse(buffer, content_type='application/pdf')
            response['Content-Disposition'] = 'attachment; filename="Tamizaje.pdf"'
            return response 
    return HttpResponseForbidden("Acceso no autorizado")

# ------------- #
# ---- FRM ---- #
# ------------- #

def datos_FRM1(request):
    query = request.GET.get("q", "").strip()

    datos_query = RespFRM.objects.select_related(
        "id_opc_frm", "id_opc_frm__id_preg_frm", "id_manychat"
    ).values(
        "id_resp_frm",
        "id_manychat",
        "id_opc_frm__id_preg_frm__preg_frm",
        "id_opc_frm__opc_resp_frm",
        "fecha_respuesta_frm",
        "id_manychat__rut_usuario",
        "id_manychat__dv_rut"
    ).order_by("-fecha_respuesta_frm")

    from .utils import filtrar_por_rut_o_manychat
    if query:
        datos_query = filtrar_por_rut_o_manychat(datos_query, query)

    page_obj = paginacion_queryset1(request, datos_query)

    return render(request, 'administracion/datos_FRM1.html', {
        "page_obj": page_obj,
        "query": query,
    })

def crear_excel_datos_frm1(request):
    session_key = request.GET.get('session_key')
    if session_key:
        session = SessionStore(session_key=session_key)
        if session.get('password_validated', False):
            session.flush()
            wb = Workbook()
            ws_FRM_V1 = wb.active
            ws_FRM_V1.title = "Factores de riesgo modificables"

            ws_FRM_V1.append(['Rut', 'Pregunta', 'Respuesta', 'Fecha Respuesta'])

            respuestas = RespFRM.objects.select_related(
                'id_opc_frm__id_preg_frm', 'id_manychat'
            ).values(
                'id_manychat__rut_usuario',
                'id_manychat__dv_rut',
                'id_opc_frm__id_preg_frm__preg_frm',
                'id_opc_frm__opc_resp_frm',
                'fecha_respuesta_frm'
            ).order_by('id_manychat__rut_usuario')

            for r in respuestas:
                fecha = r['fecha_respuesta_frm']
                ws_FRM_V1.append([
                    f"{r['id_manychat__rut_usuario']}-{r['id_manychat__dv_rut']}",
                    r['id_opc_frm__id_preg_frm__preg_frm'],
                    r['id_opc_frm__opc_resp_frm'],
                    fecha.strftime('%d-%m-%Y %H:%M:%S') if fecha else ''
                ])

            ajustar_ancho_columnas(ws_FRM_V1)
            background_colors(ws_FRM_V1)

            response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            response["Content-Disposition"] = 'attachment; filename="FactoresMod_V1.xlsx"'
            wb.save(response)
            return response
    return HttpResponseForbidden("Acceso no autorizado")

def crear_pdf_datos_frm1(request):
    session_key = request.GET.get('session_key')
    if session_key:
        session = SessionStore(session_key=session_key)
        if session.get('password_validated', False):
            session.flush()

            def truncate_text(text, max_length):
                if not text:
                    return text
                return (text[:max_length-3] + '...') if len(text) > max_length else text

            buffer = BytesIO()
            doc = SimpleDocTemplate(
                buffer,
                pagesize=landscape(A4),
                leftMargin=1*cm,
                rightMargin=1*cm,
                topMargin=1.5*cm,
                bottomMargin=1.5*cm
            )
            
            styles = getSampleStyleSheet()
            styles.add(ParagraphStyle(
                name='Small',
                parent=styles['Normal'],
                fontSize=7,
                leading=9
            ))

            respuestas = RespFRM.objects.select_related(
                'id_opc_frm__id_preg_frm', 'id_manychat'
            ).values(
                'id_manychat__rut_usuario',
                'id_manychat__dv_rut',
                'id_opc_frm__id_preg_frm__preg_frm',
                'id_opc_frm__opc_resp_frm',
                'fecha_respuesta_frm'
            ).order_by('id_manychat__rut_usuario')

            # Procesamiento de datos
            dict_respuestas = {}
            for respuesta in respuestas:
                rut = f"{respuesta['id_manychat__rut_usuario']}-{respuesta['id_manychat__dv_rut']}"
                pregunta = respuesta['id_opc_frm__id_preg_frm__preg_frm']
                respuesta_usuario = respuesta['id_opc_frm__opc_resp_frm']
                fecha = respuesta['fecha_respuesta_frm']
                
                if rut not in dict_respuestas:
                    dict_respuestas[rut] = {
                        'fecha': fecha.strftime('%d-%m-%Y %H:%M:%S') if fecha else 'Sin fecha', 
                        'respuestas': {}
                    }
                dict_respuestas[rut]['respuestas'][pregunta] = respuesta_usuario

            encabezados = ['RUT', 'Pregunta', 'Respuesta', 'Fecha Respuesta']
            data = [encabezados]

            for rut, datos in dict_respuestas.items():
                for pregunta, respuesta in datos['respuestas'].items():
                    data.append([
                        rut,
                        truncate_text(pregunta, 40),
                        truncate_text(respuesta, 30),
                        datos['fecha']
                    ])

            tabla = Table(data, repeatRows=1)
            
            # Calcular ancho de columnas
            ancho_total = landscape(A4)[0] - 2*cm  
            ancho_rut = 4*cm
            ancho_fecha = 3*cm
            ancho_pregunta = (ancho_total - ancho_rut - ancho_fecha) * 0.6
            ancho_respuesta = (ancho_total - ancho_rut - ancho_fecha) * 0.4
            
            estilo = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#c6fffa')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('FONTSIZE', (0, 1), (-1, -1), 7),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 0.25, colors.lightgrey),
                ('WORDWRAP', (0, 0), (-1, -1), True),
            ])
            
            # Aplicar anchos de columnas
            estilo.add('COLWIDTH', (0, 0), (0, -1), ancho_rut)
            estilo.add('COLWIDTH', (1, 0), (1, -1), ancho_pregunta)
            estilo.add('COLWIDTH', (2, 0), (2, -1), ancho_respuesta)
            estilo.add('COLWIDTH', (3, 0), (3, -1), ancho_fecha)
            
            # Filas alternadas
            for i in range(1, len(data)):
                if i % 2 == 0:
                    estilo.add('BACKGROUND', (0, i), (-1, i), colors.whitesmoke)
            
            tabla.setStyle(estilo)

            elementos = [
                Paragraph("Factores de Riesgo Modificables V1", styles['Title']),
                Spacer(1, 0.5*cm),
                Paragraph(f"Total de respuestas: {len(data)-1}", styles['Normal']),
                Spacer(1, 0.5*cm),
                tabla,
                Spacer(1, 0.3*cm),
                Paragraph(f"Generado el: {timezone.now().strftime('%d/%m/%Y %H:%M')}", styles['Small'])
            ]

            doc.build(elementos)
            
            buffer.seek(0)
            response = HttpResponse(buffer, content_type='application/pdf')
            response['Content-Disposition'] = 'attachment; filename="FactoresMod_V1_vertical.pdf"'
            return response
    return HttpResponseForbidden("Acceso no autorizado")

@login_required
def datos_FRM2(request):
    query = request.GET.get("q", "").strip()
    preguntas = PregFRM.objects.all()

    # QuerySet base
    usuarios_respuestas_qs = RespFRM.objects.select_related(
        "id_opc_frm", "id_opc_frm__id_preg_frm", "id_manychat"
    )

    usuarios_respuestas_qs = filtrar_por_rut_o_manychat(usuarios_respuestas_qs, query)

    usuarios_respuestas = usuarios_respuestas_qs.values(
        "id_manychat__rut_usuario",
        "id_manychat__dv_rut",
        "id_manychat",  
        "fecha_respuesta_frm",
        "id_opc_frm__id_preg_frm__preg_frm",
        "id_opc_frm__opc_resp_frm"
    )

    dict_respuestas = {}

    for respuesta in usuarios_respuestas:
        id_manychat = respuesta["id_manychat"]
        rut = respuesta["id_manychat__rut_usuario"]
        dv = respuesta["id_manychat__dv_rut"]
        rut_completo = f"{rut}-{dv}"
        pregunta = respuesta["id_opc_frm__id_preg_frm__preg_frm"]
        respuesta_usuario = respuesta["id_opc_frm__opc_resp_frm"]

        if id_manychat not in dict_respuestas:
            dict_respuestas[id_manychat] = {
                "rut_completo": rut_completo,
                "fecha": respuesta["fecha_respuesta_frm"].strftime("%d-%m-%Y %H:%M:%S"),
                "respuestas": {}
            }
        dict_respuestas[id_manychat]["respuestas"][pregunta] = respuesta_usuario

    tabla_respuestas = [
        [id_manychat, data["rut_completo"]] +
        [data["respuestas"].get(p.preg_frm, "-") for p in preguntas] +
        [data["fecha"]]
        for id_manychat, data in dict_respuestas.items()
    ]

    page_obj = paginacion_lista2(request, tabla_respuestas)

    return render(request, "administracion/datos_FRM2.html", {
        "preguntas": preguntas,
        "page_obj": page_obj,
        "query": query,
    })

def crear_excel_datos_frm2(request):
    session_key = request.GET.get('session_key')
    if session_key:
        session = SessionStore(session_key=session_key)
        if session.get('password_validated', False):
            session.flush()
            wb = Workbook()
            ws_FRM_V2 = wb.active
            ws_FRM_V2.title = "Factores de riesgo mod 2"
            
            preguntas = PregFRM.objects.all().order_by('id_preg_frm')
            lista_preguntas = ['Rut'] + [pregunta.preg_frm for pregunta in preguntas] + ['Fecha Respuesta']
            ws_FRM_V2.append(lista_preguntas)

            respuestas = RespFRM.objects.select_related(
                'id_opc_frm__id_preg_frm', 'id_manychat'
            ).values(
                'id_manychat__rut_usuario',
                'id_manychat__dv_rut',
                'id_opc_frm__id_preg_frm__preg_frm',
                'id_opc_frm__opc_resp_frm',
                'fecha_respuesta_frm'
            )

            dict_respuestas = {}
            for respuesta in respuestas:
                rut = f"{respuesta['id_manychat__rut_usuario']}-{respuesta['id_manychat__dv_rut']}"
                pregunta = respuesta['id_opc_frm__id_preg_frm__preg_frm']
                respuesta_usuario = respuesta['id_opc_frm__opc_resp_frm']
                fecha = respuesta['fecha_respuesta_frm'].strftime("%d-%m-%Y %H:%M:%S") if respuesta['fecha_respuesta_frm'] else ''
                
                if rut not in dict_respuestas:
                    dict_respuestas[rut] = {
                        "respuestas": {},
                        "fecha": fecha
                    }
                dict_respuestas[rut]["respuestas"][pregunta] = respuesta_usuario

            for rut, respuestas_usuario in dict_respuestas.items():
                fila = [rut]
                for pregunta in preguntas:
                    respuesta = respuestas_usuario["respuestas"].get(pregunta.preg_frm, '')
                    fila.append(respuesta)
                fila.append(respuestas_usuario["fecha"])
                ws_FRM_V2.append(fila)
        
            ajustar_ancho_columnas(ws_FRM_V2)
            background_colors(ws_FRM_V2)

            response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            response["Content-Disposition"] = 'attachment; filename="FactoresMod_V2.xlsx"'
            wb.save(response)
            return response
    return HttpResponseForbidden("Acceso no autorizado")

@login_required
def crear_pdf_datos_frm2(request):
    session_key = request.GET.get('session_key')
    if session_key:
        session = SessionStore(session_key=session_key)
        if session.get('password_validated', False):
            session.flush()
            def truncate_text(text, max_length):
                if not text:
                    return text
                return (text[:max_length-3] + '...') if len(text) > max_length else text

        
            preguntas = PregFRM.objects.all().order_by('id_preg_frm')
            
            respuestas = RespFRM.objects.select_related(
                'id_opc_frm__id_preg_frm', 'id_manychat'
            ).values(
                'id_manychat__rut_usuario',
                'id_manychat__dv_rut',
                'id_opc_frm__id_preg_frm__preg_frm',
                'id_opc_frm__opc_resp_frm',
                'fecha_respuesta_frm'
            )

            dict_respuestas = {}
            for respuesta in respuestas:
                rut = f"{respuesta['id_manychat__rut_usuario']}-{respuesta['id_manychat__dv_rut']}"
                pregunta = respuesta['id_opc_frm__id_preg_frm__preg_frm']
                respuesta_usuario = respuesta['id_opc_frm__opc_resp_frm']
                fecha = respuesta['fecha_respuesta_frm']
                
                if rut not in dict_respuestas:
                    dict_respuestas[rut] = {
                        'fecha': fecha.strftime('%d-%m-%Y %H:%M:%S') if fecha else 'Sin fecha',
                        'respuestas': {}
                    }
                dict_respuestas[rut]['respuestas'][pregunta] = respuesta_usuario

            buffer = BytesIO()
            doc = SimpleDocTemplate(
                buffer,
                pagesize=landscape(A4),
                leftMargin=1*cm,
                rightMargin=1*cm,
                topMargin=1.5*cm,
                bottomMargin=1.5*cm
            )
            
            styles = getSampleStyleSheet()
            styles.add(ParagraphStyle(
                name='Small',
                parent=styles['Normal'],
                fontSize=7,
                leading=9
            ))

            encabezados = ['RUT'] + [truncate_text(p.preg_frm, 25) for p in preguntas] + ['Fecha Respuesta']
            data = [encabezados]

            for rut, datos in dict_respuestas.items():
                fila = [rut]
                for p in preguntas:
                    respuesta = datos['respuestas'].get(p.preg_frm, 'NR') 
                    fila.append(truncate_text(respuesta, 20))
                fila.append(datos['fecha'])
                data.append(fila)


            tabla = Table(data, repeatRows=1)
            
            ancho_total = landscape(A4)[0] - 2*cm  
            ancho_rut = 6*cm
            ancho_fecha = 4*cm
            ancho_preguntas = max(3*cm, (ancho_total - ancho_rut - ancho_fecha) / len(preguntas))
            
            estilo = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#c6fffa')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 7),
                ('FONTSIZE', (0, 1), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 0.25, colors.lightgrey),
                ('WORDWRAP', (0, 0), (-1, -1), True),
            ])
            
            estilo.add('COLWIDTH', (0, 0), (0, -1), ancho_rut)
            for i in range(1, len(preguntas)+1):
                estilo.add('COLWIDTH', (i, 0), (i, -1), ancho_preguntas)
            estilo.add('COLWIDTH', (-1, 0), (-1, -1), ancho_fecha)
            
            
            for i in range(1, len(data)):
                if i % 2 == 0:
                    estilo.add('BACKGROUND', (0, i), (-1, i), colors.whitesmoke)
            
            tabla.setStyle(estilo)

        
            elementos = [
                Paragraph("Factores de Riesgo Modificables V2", styles['Title']),
                Spacer(1, 0.5*cm),
                Paragraph(f"Total de registros: {len(data)-1}", styles['Normal']),
                Spacer(1, 0.5*cm),
                tabla,
                Spacer(1, 0.3*cm),
                Paragraph(f"Generado el: {timezone.now().strftime('%d/%m/%Y %H:%M')} | NR = No Respondió", styles['Small'])
            ]

            doc.build(elementos)
            buffer.seek(0)
            response = HttpResponse(buffer, content_type='application/pdf')
            response['Content-Disposition'] = 'attachment; filename="FactoresMod_V2.pdf"'
            return response
    return HttpResponseForbidden("Acceso no autorizado")

# -------------- #
# ---- FRNM ---- #
# -------------- #

@login_required
def datos_FRNM1(request):
    query = request.GET.get("q", "").strip()

    datos_qs = RespFRNM.objects.select_related(
        "id_opc_frnm", "id_opc_frnm__id_preg_frnm", "id_manychat"
    )

    datos_qs = filtrar_por_rut_o_manychat(datos_qs, query)

    datos_query = datos_qs.values(
        "id_resp_frnm",
        "id_manychat",
        "id_opc_frnm__id_preg_frnm__preg_frnm",
        "id_opc_frnm__opc_resp_frnm",
        "fecha_respuesta_frnm",
        "id_manychat__rut_usuario",
        "id_manychat__dv_rut"
    ).order_by("-fecha_respuesta_frnm")

    page_obj = paginacion_queryset1(request, datos_query) 

    return render(request, 'administracion/datos_FRNM1.html', {
        "page_obj": page_obj,
        "query": query,
    })

def crear_excel_datos_frnm1(request):
    session_key = request.GET.get('session_key')
    if session_key:
        session = SessionStore(session_key=session_key)
        if session.get('password_validated', False):
            session.flush()
            wb = Workbook()
            ws_FRNM_V1 = wb.active
            ws_FRNM_V1.title = "Riesgos no modificables"

            ws_FRNM_V1.append(['Rut', 'Pregunta', 'Respuesta', 'Fecha Respuesta'])

            respuestas = RespFRNM.objects.select_related(
                'id_opc_frnm__id_preg_frnm', 'id_manychat'
            ).values(
                'id_manychat__rut_usuario',
                'id_manychat__dv_rut',
                'id_opc_frnm__id_preg_frnm__preg_frnm',
                'id_opc_frnm__opc_resp_frnm',
                'fecha_respuesta_frnm'
            ).order_by('id_manychat__rut_usuario')

            for r in respuestas:
                fecha = r['fecha_respuesta_frnm']
                ws_FRNM_V1.append([
                    f"{r['id_manychat__rut_usuario']}-{r['id_manychat__dv_rut']}",
                    r['id_opc_frnm__id_preg_frnm__preg_frnm'],
                    r['id_opc_frnm__opc_resp_frnm'],
                    fecha.strftime('%d-%m-%Y %H:%M:%S') if fecha else ''
                ])

            ajustar_ancho_columnas(ws_FRNM_V1)
            background_colors(ws_FRNM_V1)

            response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            response["Content-Disposition"] = 'attachment; filename="FactoresNoMod_V1.xlsx"'
            wb.save(response)
            return response
    return HttpResponseForbidden("Acceso no autorizado")

@login_required
def crear_pdf_datos_frnm1(request):
    session_key = request.GET.get('session_key')
    if session_key:
        session = SessionStore(session_key=session_key)
        if session.get('password_validated', False):
            session.flush()
            respuestas = RespFRNM.objects.select_related(
                'id_opc_frnm__id_preg_frnm', 'id_manychat'
            ).values(
                'id_manychat__rut_usuario',
                'id_manychat__dv_rut',
                'id_opc_frnm__id_preg_frnm__preg_frnm',
                'id_opc_frnm__opc_resp_frnm',
                'fecha_respuesta_frnm'
            ).order_by('-fecha_respuesta_frnm')

            buffer = BytesIO()
            doc = SimpleDocTemplate(
                buffer,
                pagesize=letter,
                leftMargin=1.5*cm,
                rightMargin=1.5*cm,
                topMargin=2*cm,
                bottomMargin=2*cm
            )
            
            styles = getSampleStyleSheet()
            if 'Small' not in styles:
                styles.add(ParagraphStyle(
                    name='Small',
                    parent=styles['Normal'],
                    fontSize=8,
                    leading=10
                ))

            data = []
            encabezados = ['RUT', 'Pregunta', 'Respuesta', 'Fecha Respuesta']
            data.append(encabezados)

            for r in respuestas:
                fecha = r['fecha_respuesta_frnm']
                data.append([
                    f"{r['id_manychat__rut_usuario']}-{r['id_manychat__dv_rut']}",
                    r['id_opc_frnm__id_preg_frnm__preg_frnm'],
                    r['id_opc_frnm__opc_resp_frnm'],
                    fecha.strftime('%d-%m-%Y %H:%M:%S') if fecha else 'Sin fecha'
                ])

            tabla = Table(data, repeatRows=1)
            
            estilo = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#c6fffa')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('WORDWRAP', (0, 0), (-1, -1), True),
            ])
            
            ancho_columnas = [7*cm, 8*cm, 6*cm, 4*cm]
            for i, width in enumerate(ancho_columnas):
                estilo.add('COLWIDTH', (i, 0), (i, -1), width)
            
            for i in range(1, len(data)):
                if i % 2 == 0:
                    estilo.add('BACKGROUND', (0, i), (-1, i), colors.whitesmoke)
            
            tabla.setStyle(estilo)

            elementos = [
                Paragraph("Factores de Riesgo No Modificables V1", styles['Title']),
                Spacer(1, 0.5*cm),
                Paragraph(f"Total de registros: {len(data)-1}", styles['Normal']),
                Spacer(1, 0.5*cm),
                tabla,
                Spacer(1, 0.3*cm),
                Paragraph("Generado el: " + timezone.now().strftime('%d/%m/%Y %H:%M'), styles['Small'])  
            ]

            doc.build(elementos)
            
            buffer.seek(0)
            response = HttpResponse(buffer, content_type='application/pdf')
            response['Content-Disposition'] = 'attachment; filename="FactoresNoMod_V1.pdf"'
            return response
    return HttpResponseForbidden("Acceso no autorizado")

@login_required
def datos_FRNM2(request):
    query = request.GET.get("q", "").strip()

    preguntas = PregFRNM.objects.all()

    usuarios_respuestas_qs = RespFRNM.objects.select_related(
        "id_opc_frnm", "id_opc_frnm__id_preg_frnm", "id_manychat"
    )

    usuarios_respuestas_qs = filtrar_por_rut_o_manychat(usuarios_respuestas_qs, query)

    usuarios_respuestas = usuarios_respuestas_qs.values(
        "id_manychat__rut_usuario",
        "id_manychat__dv_rut",
        "id_manychat",
        "fecha_respuesta_frnm",
        "id_opc_frnm__id_preg_frnm__preg_frnm",
        "id_opc_frnm__opc_resp_frnm"
    )

    dict_respuestas = {}

    for respuesta in usuarios_respuestas:
        id_manychat = respuesta["id_manychat"]
        rut = respuesta["id_manychat__rut_usuario"]
        dv = respuesta["id_manychat__dv_rut"]
        rut_completo = f"{rut}-{dv}"
        pregunta = respuesta["id_opc_frnm__id_preg_frnm__preg_frnm"]
        respuesta_usuario = respuesta["id_opc_frnm__opc_resp_frnm"]

        if id_manychat not in dict_respuestas:
            dict_respuestas[id_manychat] = {
                "rut_completo": rut_completo,
                "fecha": respuesta["fecha_respuesta_frnm"].strftime("%d-%m-%Y %H:%M:%S"),
                "respuestas": {}
            }
        dict_respuestas[id_manychat]["respuestas"][pregunta] = respuesta_usuario

    tabla_respuestas = [
        [id_manychat, data["rut_completo"]] + [data["respuestas"].get(p.preg_frnm, "-") for p in preguntas] + [data["fecha"]]
        for id_manychat, data in dict_respuestas.items()
    ]

    page_obj = paginacion_lista2(request, tabla_respuestas)

    return render(request, "administracion/datos_FRNM2.html", {
        "preguntas": preguntas,
        "page_obj": page_obj,
        "query": query,
    })

def crear_excel_datos_frnm2(request):
    session_key = request.GET.get('session_key')
    if session_key:
        session = SessionStore(session_key=session_key)
        if session.get('password_validated', False):
            session.flush()
            wb = Workbook()
            ws_FRNM_V2 = wb.active
            ws_FRNM_V2.title = "Factores de riesgo no mod 2"
            
            preguntas = PregFRNM.objects.all().order_by('id_preg_frnm')
            lista_preguntas = ['Rut'] + [pregunta.preg_frnm for pregunta in preguntas] + ['Fecha Respuesta']
            ws_FRNM_V2.append(lista_preguntas)

            respuestas = RespFRNM.objects.select_related(
                'id_opc_frnm__id_preg_frnm', 'id_manychat'
            ).values(
                'id_manychat__rut_usuario',
                'id_manychat__dv_rut',
                'id_opc_frnm__id_preg_frnm__preg_frnm',
                'id_opc_frnm__opc_resp_frnm',
                'fecha_respuesta_frnm'
            )

            dict_respuestas = {}
            for respuesta in respuestas:
                rut = f"{respuesta['id_manychat__rut_usuario']}-{respuesta['id_manychat__dv_rut']}"
                pregunta = respuesta['id_opc_frnm__id_preg_frnm__preg_frnm']
                respuesta_usuario = respuesta['id_opc_frnm__opc_resp_frnm']
                fecha = respuesta['fecha_respuesta_frnm'].strftime('%d-%m-%Y %H-%M-%S')if respuesta['fecha_respuesta_frnm'] else ''
                
                if rut not in dict_respuestas:
                    dict_respuestas[rut] = {
                        "respuestas": {},
                        "fecha": fecha
                    }
                dict_respuestas[rut]["respuestas"][pregunta] = respuesta_usuario

            for rut, respuestas_usuario in dict_respuestas.items():
                fila = [rut]
                for pregunta in preguntas:
                    respuesta = respuestas_usuario["respuestas"].get(pregunta.preg_frnm, '')
                    fila.append(respuesta)
                fila.append(respuestas_usuario["fecha"])
                ws_FRNM_V2.append(fila)
        
            ajustar_ancho_columnas(ws_FRNM_V2)
            background_colors(ws_FRNM_V2)

            response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            response["Content-Disposition"] = 'attachment; filename="FactoresNoMod_V2.xlsx"'
            wb.save(response)
            return response
    return HttpResponseForbidden("Acceso no autorizado")

@login_required
def crear_pdf_datos_frnm2(request):
    session_key = request.GET.get('session_key')
    if session_key:
        session = SessionStore(session_key=session_key)
        if session.get('password_validated', False):
            session.flush()
            def truncate_text(text, max_length):
                if not text:
                    return text
                return (text[:max_length-3] + '...') if len(text) > max_length else text

            preguntas = PregFRNM.objects.all().order_by('id_preg_frnm')
            
            respuestas = RespFRNM.objects.select_related(
                'id_opc_frnm__id_preg_frnm', 'id_manychat'
            ).values(
                'id_manychat__rut_usuario',
                'id_manychat__dv_rut',
                'id_opc_frnm__id_preg_frnm__preg_frnm',
                'id_opc_frnm__opc_resp_frnm',
                'fecha_respuesta_frnm'
            )

            dict_respuestas = {}
            for respuesta in respuestas:
                rut = f"{respuesta['id_manychat__rut_usuario']}-{respuesta['id_manychat__dv_rut']}"
                pregunta = respuesta['id_opc_frnm__id_preg_frnm__preg_frnm']
                respuesta_usuario = respuesta['id_opc_frnm__opc_resp_frnm']
                fecha = respuesta['fecha_respuesta_frnm']
                
                if rut not in dict_respuestas:
                    dict_respuestas[rut] = {'fecha': fecha, 'respuestas': {}}
                dict_respuestas[rut]['respuestas'][pregunta] = respuesta_usuario

            buffer = BytesIO()
            doc = SimpleDocTemplate(
                buffer,
                pagesize=landscape(A4),
                leftMargin=1*cm,
                rightMargin=1*cm,
                topMargin=1.5*cm,
                bottomMargin=1.5*cm
            )
            
            styles = getSampleStyleSheet()
            styles.add(ParagraphStyle(
                name='SmallText',
                parent=styles['Normal'],
                fontSize=6,
                leading=8
            ))

            encabezados = ['RUT'] + [truncate_text(p.preg_frnm, 25) for p in preguntas] + ['Fecha Respuesta']
            data = [encabezados]

            for rut, datos in dict_respuestas.items():
                fila = [rut]
                for p in preguntas:
                    respuesta = datos['respuestas'].get(p.preg_frnm, 'NR')  
                    fila.append(truncate_text(respuesta, 20))
                fila.append(datos['fecha'].strftime('%d-%m-%Y %H:%M:%S') if datos['fecha'] else 'S/F')
                data.append(fila)

            tabla = Table(data, repeatRows=1)
            
            ancho_total = landscape(A4)[0] - 2*cm 
            ancho_rut = 6*cm
            ancho_fecha = 3*cm
            ancho_preguntas = max(2.5*cm, (ancho_total - ancho_rut - ancho_fecha) / len(preguntas))
            
            estilo = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#c6fffa')), 
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 7),
                ('FONTSIZE', (0, 1), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 0.25, colors.lightgrey),
                ('WORDWRAP', (0, 0), (-1, -1), True),  
                ('LEADING', (0, 0), (-1, -1), 7), 
            ])
            
            estilo.add('COLWIDTH', (0, 0), (0, -1), ancho_rut) 
            for i in range(1, len(preguntas)+1):
                estilo.add('COLWIDTH', (i, 0), (i, -1), ancho_preguntas)  
            estilo.add('COLWIDTH', (-1, 0), (-1, -1), ancho_fecha)  
            
            for i in range(1, len(data)):
                if i % 2 == 0:
                    estilo.add('BACKGROUND', (0, i), (-1, i), colors.whitesmoke)
            
            tabla.setStyle(estilo)

            elementos = [
                Paragraph("Factores de Riesgo No Modificables V2", styles['Title']),
                Spacer(1, 0.5*cm),
                Paragraph(f"Total de registros: {len(data)-1}", styles['Normal']),
                Spacer(1, 0.3*cm),
                tabla,
                Spacer(1, 0.3*cm),
                Paragraph("NR = No Respondió | S/F = Sin Fecha", styles['SmallText'])
            ]

            doc.build(elementos)
            
            buffer.seek(0)
            response = HttpResponse(buffer, content_type='application/pdf')
            response['Content-Disposition'] = 'attachment; filename="FactoresRiesgoNoMod_V2.pdf"'
            return response
    return HttpResponseForbidden("Acceso no autorizado")

# ------------ #
# ---- DS ---- #
# ------------ #

@login_required
def datos_DS1(request):
    query = request.GET.get("q", "").strip()

    datos_qs = RespDS.objects.select_related(
        "id_opc_ds", "id_opc_ds__id_preg_ds", "id_manychat"
    )

    datos_qs = filtrar_por_rut_o_manychat(datos_qs, query)

    datos_query = datos_qs.values(
        "id_resp_ds",
        "id_manychat",
        "id_opc_ds__id_preg_ds__preg_ds",
        "id_opc_ds__opc_resp_ds",
        "fecha_respuesta_ds",
        "id_manychat__rut_usuario",
        "id_manychat__dv_rut"
    ).order_by("-fecha_respuesta_ds")

    page_obj = paginacion_queryset1(request, datos_query) 

    return render(request, 'administracion/datos_DS1.html', {
        "page_obj": page_obj,
        "query": query,
    })

def crear_excel_datos_ds1(request):
    session_key = request.GET.get('session_key')
    if session_key:
        session = SessionStore(session_key=session_key)
        if session.get('password_validated', False):
            session.flush()
            wb = Workbook()
            ws_DS_V1 = wb.active
            ws_DS_V1.title = "Determinantes sociales"

            ws_DS_V1.append(['Rut', 'Pregunta', 'Respuesta', 'Fecha Respuesta'])

            respuestas = RespDS.objects.select_related(
                'id_opc_ds__id_preg_ds', 'id_manychat'
            ).values(
                'id_manychat__rut_usuario',
                'id_manychat__dv_rut',
                'id_opc_ds__id_preg_ds__preg_ds',
                'id_opc_ds__opc_resp_ds',
                'fecha_respuesta_ds'
            ).order_by('id_manychat__rut_usuario')

            for r in respuestas:
                fecha = r['fecha_respuesta_ds']
                ws_DS_V1.append([
                    f"{r['id_manychat__rut_usuario']}-{r['id_manychat__dv_rut']}",
                    r['id_opc_ds__id_preg_ds__preg_ds'],
                    r['id_opc_ds__opc_resp_ds'],
                    fecha.strftime('%d-%m-%Y %H:%M:%S') if fecha else ''
                ])
            
            ajustar_ancho_columnas(ws_DS_V1)
            background_colors(ws_DS_V1)

            response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            response["Content-Disposition"] = 'attachment; filename="DeterminantesSociales_V1.xlsx"'
            wb.save(response)
            return response
    return HttpResponseForbidden("Acceso no autorizado")

@login_required
def crear_pdf_datos_ds1(request):
    session_key = request.GET.get('session_key')
    if session_key:
        session = SessionStore(session_key=session_key)
        if session.get('password_validated', False):
            session.flush()
            respuestas = RespDS.objects.select_related(
                'id_opc_ds__id_preg_ds', 'id_manychat'
            ).values(
                'id_manychat__rut_usuario',
                'id_manychat__dv_rut',
                'id_opc_ds__id_preg_ds__preg_ds',
                'id_opc_ds__opc_resp_ds',
                'fecha_respuesta_ds'
            ).order_by('-fecha_respuesta_ds') 

            data = []
            
            encabezados = ['RUT', 'Pregunta', 'Respuesta', 'Fecha Respuesta']
            data.append(encabezados)

            for r in respuestas:
                fecha = r['fecha_respuesta_ds']
                data.append([
                    f"{r['id_manychat__rut_usuario']}-{r['id_manychat__dv_rut']}",
                    r['id_opc_ds__id_preg_ds__preg_ds'],
                    r['id_opc_ds__opc_resp_ds'],
                    fecha.strftime('%d-%m-%Y %H:%M:%S') if fecha else 'Sin fecha'
                ])

            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=letter)
            styles = getSampleStyleSheet()
            
            tabla = Table(data)
            
            estilo = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#c6fffa')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
            ])
            
            for i in range(1, len(data)):
                if i % 2 == 0:
                    estilo.add('BACKGROUND', (0, i), (-1, i), colors.whitesmoke)
            
            ancho_columnas = [100, 200, 150, 100] 
            for i, width in enumerate(ancho_columnas):
                estilo.add('COLWIDTH', (i, 0), (i, -1), width)
            
            tabla.setStyle(estilo)
            
            elementos = []
            
            titulo = Paragraph("Determinantes Sociales V1", styles['Title'])
            elementos.append(titulo)
            
            elementos.append(Paragraph("<br/><br/>", styles['Normal']))
            
            elementos.append(tabla)
            
            doc.build(elementos)
            
            buffer.seek(0)
            response = HttpResponse(buffer, content_type='application/pdf')
            response['Content-Disposition'] = 'attachment; filename="DeterminantesSociales_V1.pdf"'
            
            return response
    return HttpResponseForbidden("Acceso no autorizado")

@login_required
def datos_DS2(request):
    query = request.GET.get("q", "").strip()

    preguntas = PregDS.objects.all()
    
    usuarios_respuestas_qs = RespDS.objects.select_related(
        "id_opc_ds", "id_opc_ds__id_preg_ds", "id_manychat"
    )

    usuarios_respuestas_qs = filtrar_por_rut_o_manychat(usuarios_respuestas_qs, query)

    usuarios_respuestas = usuarios_respuestas_qs.values(
        "id_manychat__rut_usuario",
        "id_manychat__dv_rut",
        "id_manychat", 
        "fecha_respuesta_ds",
        "id_opc_ds__id_preg_ds__preg_ds",
        "id_opc_ds__opc_resp_ds"
    )

    dict_respuestas = {}

    for respuesta in usuarios_respuestas:
        id_manychat = respuesta["id_manychat"]
        rut = respuesta["id_manychat__rut_usuario"]
        dv = respuesta["id_manychat__dv_rut"]
        rut_completo = f"{rut}-{dv}"
        pregunta = respuesta["id_opc_ds__id_preg_ds__preg_ds"]
        respuesta_usuario = respuesta["id_opc_ds__opc_resp_ds"]
        
        if id_manychat not in dict_respuestas:
            dict_respuestas[id_manychat] = {
                "rut_completo": rut_completo,
                "fecha": respuesta["fecha_respuesta_ds"].strftime('%d-%m-%Y %H:%M:%S'),
                "respuestas": {}
            }
        dict_respuestas[id_manychat]["respuestas"][pregunta] = respuesta_usuario

    tabla_respuestas = [
        [id_manychat, data["rut_completo"]] + [data["respuestas"].get(p.preg_ds, "-") for p in preguntas] + [data["fecha"]]
        for id_manychat, data in dict_respuestas.items()
    ]

    page_obj = paginacion_lista2(request, tabla_respuestas) 

    return render(request, "administracion/datos_DS2.html", {
        "preguntas": preguntas,
        "page_obj": page_obj,
        "query": query,
    })

def crear_excel_datos_ds2(request):
    session_key = request.GET.get('session_key')
    if session_key:
        session = SessionStore(session_key=session_key)
        if session.get('password_validated', False):
            session.flush()
            wb = Workbook()
            ws_DS_V2 = wb.active
            ws_DS_V2.title = "Determinantes salud 2"
            
            preguntas = PregDS.objects.all().order_by('id_preg_ds')
            lista_preguntas = ['Rut'] + [pregunta.preg_ds for pregunta in preguntas] + ['Fecha Respuesta']
            ws_DS_V2.append(lista_preguntas)

            respuestas = RespDS.objects.select_related(
                'id_opc_ds__id_preg_ds', 'id_manychat'
            ).values(
                'id_manychat__rut_usuario',
                'id_manychat__dv_rut',
                'id_opc_ds__id_preg_ds__preg_ds',
                'id_opc_ds__opc_resp_ds',
                'fecha_respuesta_ds'
            )

            dict_respuestas = {}
            for respuesta in respuestas:
                rut = f"{respuesta['id_manychat__rut_usuario']}-{respuesta['id_manychat__dv_rut']}"
                pregunta = respuesta['id_opc_ds__id_preg_ds__preg_ds']
                respuesta_usuario = respuesta['id_opc_ds__opc_resp_ds']
                fecha = respuesta['fecha_respuesta_ds'].strftime("%d-%m-%Y %H:%M:%S") if respuesta['fecha_respuesta_ds'] else ''
                
                if rut not in dict_respuestas:
                    dict_respuestas[rut] = {
                        "respuestas": {},
                        "fecha": fecha
                    }
                dict_respuestas[rut]["respuestas"][pregunta] = respuesta_usuario

            for rut, respuestas_usuario in dict_respuestas.items():
                fila = [rut]
                for pregunta in preguntas:
                    respuesta = respuestas_usuario["respuestas"].get(pregunta.preg_ds, '')
                    fila.append(respuesta)
                fila.append(respuestas_usuario["fecha"])
                ws_DS_V2.append(fila)
        
            ajustar_ancho_columnas(ws_DS_V2)
            background_colors(ws_DS_V2)

            response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            response["Content-Disposition"] = 'attachment; filename="DeterminantesSalud_V2.xlsx"'
            wb.save(response)
            return response
    return HttpResponseForbidden("Acceso no autorizado")

@login_required
def crear_pdf_datos_ds2(request):
    session_key = request.GET.get('session_key')
    if session_key:
        session = SessionStore(session_key=session_key)
        if session.get('password_validated', False):
            session.flush()
            def truncate_text(text, max_length):
                if not text:
                    return text
                return (text[:max_length-3] + '...') if len(text) > max_length else text

        
            preguntas = PregDS.objects.all().order_by('id_preg_ds')
            
            respuestas = RespDS.objects.select_related(
                'id_opc_ds__id_preg_ds', 'id_manychat'
            ).values(
                'id_manychat__rut_usuario',
                'id_manychat__dv_rut',
                'id_opc_ds__id_preg_ds__preg_ds',
                'id_opc_ds__opc_resp_ds',
                'fecha_respuesta_ds'
            )

            dict_respuestas = {}
            for respuesta in respuestas:
                rut = f"{respuesta['id_manychat__rut_usuario']}-{respuesta['id_manychat__dv_rut']}"
                pregunta = respuesta['id_opc_ds__id_preg_ds__preg_ds']
                respuesta_usuario = respuesta['id_opc_ds__opc_resp_ds']
                fecha = respuesta['fecha_respuesta_ds']
                
                if rut not in dict_respuestas:
                    dict_respuestas[rut] = {
                        'fecha': fecha.strftime('%d-%m-%Y %H:%M:%S') if fecha else 'Sin fecha',
                        'respuestas': {}
                    }
                dict_respuestas[rut]['respuestas'][pregunta] = respuesta_usuario

            buffer = BytesIO()
            doc = SimpleDocTemplate(
                buffer,
                pagesize=landscape(A4),
                leftMargin=1*cm,
                rightMargin=1*cm,
                topMargin=1.5*cm,
                bottomMargin=1.5*cm
            )
            
            styles = getSampleStyleSheet()
            styles.add(ParagraphStyle(
                name='Small',
                parent=styles['Normal'],
                fontSize=7,
                leading=9
            ))

            encabezados = ['RUT'] + [truncate_text(p.preg_ds, 25) for p in preguntas] + ['Fecha Respuesta']
            data = [encabezados]

            for rut, datos in dict_respuestas.items():
                fila = [rut]
                for p in preguntas:
                    respuesta = datos['respuestas'].get(p.preg_ds, 'NR') 
                    fila.append(truncate_text(respuesta, 20))
                fila.append(datos['fecha'])
                data.append(fila)

            tabla = Table(data, repeatRows=1)
            
            ancho_total = landscape(A4)[0] - 2*cm  
            ancho_rut = 6*cm
            ancho_fecha = 4*cm
            ancho_preguntas = max(3*cm, (ancho_total - ancho_rut - ancho_fecha) / len(preguntas))
            
            estilo = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#c6fffa')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 7),
                ('FONTSIZE', (0, 1), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 0.25, colors.lightgrey),
                ('WORDWRAP', (0, 0), (-1, -1), True),
            ])
            
            estilo.add('COLWIDTH', (0, 0), (0, -1), ancho_rut)
            for i in range(1, len(preguntas)+1):
                estilo.add('COLWIDTH', (i, 0), (i, -1), ancho_preguntas)
            estilo.add('COLWIDTH', (-1, 0), (-1, -1), ancho_fecha)
            
            
            for i in range(1, len(data)):
                if i % 2 == 0:
                    estilo.add('BACKGROUND', (0, i), (-1, i), colors.whitesmoke)
            
            tabla.setStyle(estilo)

        
            elementos = [
                Paragraph("Determinantes de Salud V2", styles['Title']),
                Spacer(1, 0.5*cm),
                Paragraph(f"Total de registros: {len(data)-1}", styles['Normal']),
                Spacer(1, 0.5*cm),
                tabla,
                Spacer(1, 0.3*cm),
                Paragraph(f"Generado el: {timezone.now().strftime('%d/%m/%Y %H:%M')} | NR = No Respondió", styles['Small'])
            ]

            doc.build(elementos)
            buffer.seek(0)
            response = HttpResponse(buffer, content_type='application/pdf')
            response['Content-Disposition'] = 'attachment; filename="DeterminantesSalud_V2.pdf"'
            return response
    return HttpResponseForbidden("Acceso no autorizado")
    
# ----------------------------------------------------------------- #
# ---------------------- Listado priorizado ----------------------- #
# ----------------------------------------------------------------- #
@never_cache
@login_required
def listado_priorizado(request):
    query = request.GET.get("q", "").strip()

    if request.method == "POST":
        password_ingresada = request.POST.get("password", "").strip()
        if password_ingresada == settings.ACCESO_LISTADO:
            request.session["acceso_listado_permitido"] = True
            return redirect("listado_priorizado")
        else:
            return render(request, "administracion/form_contrasena_listado.html", {
                "error": "Contraseña incorrecta"
            })

    # Si no hay permiso de sesión: pedir contraseña
    if not request.session.get("acceso_listado_permitido", False):
        return render(request, "administracion/form_contrasena_listado.html")

    # Acceso permitido: mostrar listado
    pap_subquery = RespTM.objects.filter(
        id_manychat=OuterRef('id_manychat'),
        id_opc_tm=5
    ).values('id_opc_tm')[:1]

    parejas_subquery = RespFRNM.objects.filter(
        id_manychat=OuterRef('id_manychat'),
        id_opc_frnm=9
    ).values('id_opc_frnm')[:1]

    usuarios = Usuario.objects.annotate(
        nombre_comuna=F('cod_comuna__nombre_comuna'),
        tiene_pap=Subquery(pap_subquery),
        tiene_parejas=Subquery(parejas_subquery),
        pap_alterado=Case(
            When(tiene_pap__isnull=False, then=Value('Sí')),
            default=Value('No aplica'),
            output_field=CharField()
        ),
        parejas_sexuales=Case(
            When(tiene_parejas__isnull=False, then=Value('Sí')),
            default=Value('No aplica'),
            output_field=CharField()
        )
    ).order_by("id_manychat")

    if query:
        usuarios = filtro_listado_priorizado(usuarios, query)

    datos_procesados = []
    for usuario in usuarios:
        edad = calcular_edad(usuario.fecha_nacimiento) if usuario.fecha_nacimiento else None
        datos_procesados.append({
            "id": usuario.id_manychat,
            "rut_usuario": f"{usuario.rut_usuario}-{usuario.dv_rut}",
            "dv_rut": usuario.dv_rut,
            "num_whatsapp": usuario.num_whatsapp,
            "email": usuario.email or "No disponible",
            "edad": edad,
            "nombre_comuna": usuario.nombre_comuna,
            "pap_alterado": usuario.pap_alterado,
            "parejas_sexuales": usuario.parejas_sexuales
        })

    paginator = Paginator(datos_procesados, 20)
    page_number = request.GET.get("page")
    try:
        page_obj = paginator.page(page_number)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)

    return render(request, "administracion/listado_priorizado.html", {
        "page_obj": page_obj,
        "query": query
    })


@login_required
def salir_listado_priorizado(request):
    request.session.pop("acceso_listado_permitido", None) 
    return redirect("respuestas")

def crear_excel_listado_priorizado(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Listado Priorizado"

    # Encabezados
    encabezados = [
        'ID ManyChat', 
        'RUT', 
        'WhatsApp', 
        'Email', 
        'Edad',
        'Comuna', 
        'PAP Alterado', 
        'Parejas sexuales'
    ]
    ws.append(encabezados)

    # Obtener los datos
    pap_subquery = RespTM.objects.filter(
        id_manychat=OuterRef('id_manychat'),
        id_opc_tm=5
    ).values('id_opc_tm')[:1]

    parejas_subquery = RespFRNM.objects.filter(
        id_manychat=OuterRef('id_manychat'),
        id_opc_frnm=9
    ).values('id_opc_frnm')[:1]

    usuarios = Usuario.objects.annotate(
        nombre_comuna=F('cod_comuna__nombre_comuna'),
        tiene_pap=Subquery(pap_subquery),
        tiene_parejas=Subquery(parejas_subquery),
        pap_alterado=Case(
            When(tiene_pap__isnull=False, then=Value('Sí')),
            default=Value('No aplica'),
            output_field=CharField()
        ),
        parejas_sexuales=Case(
            When(tiene_parejas__isnull=False, then=Value('Sí')),
            default=Value('No aplica'),
            output_field=CharField()
        )
    ).order_by('id_manychat')

    # Llenar el Excel con los datos
    for usuario in usuarios:
        edad = calcular_edad(usuario.fecha_nacimiento) if usuario.fecha_nacimiento else None
        
        fila = [
            usuario.id_manychat,
            f"{usuario.rut_usuario}-{usuario.dv_rut}",
            usuario.num_whatsapp,
            usuario.email or 'No disponible',
            edad or 'No registrada',
            usuario.nombre_comuna,
            usuario.pap_alterado,
            usuario.parejas_sexuales
        ]
        ws.append(fila)
    
    # Ajustar formato
    ajustar_ancho_columnas(ws)
    background_colors(ws)

    # Preparar la respuesta
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="Listado_Priorizado.xlsx"'

    wb.save(response)
    return response

@login_required
def crear_pdf_listado_priorizado(request):
    pap_subquery = RespTM.objects.filter(
        id_manychat=OuterRef('id_manychat'),
        id_opc_tm=5
    ).values('id_opc_tm')[:1]

    parejas_subquery = RespFRNM.objects.filter(
        id_manychat=OuterRef('id_manychat'),
        id_opc_frnm=9
    ).values('id_opc_frnm')[:1]

    usuarios = Usuario.objects.annotate(
        nombre_comuna=F('cod_comuna__nombre_comuna'),
        tiene_pap=Subquery(pap_subquery),
        tiene_parejas=Subquery(parejas_subquery),
        pap_alterado=Case(
            When(tiene_pap__isnull=False, then=Value('Sí')),
            default=Value('No aplica'),
            output_field=CharField()
        ),
        parejas_sexuales=Case(
            When(tiene_parejas__isnull=False, then=Value('Sí')),
            default=Value('No aplica'),
            output_field=CharField()
        )
    ).order_by('id_manychat')

    data = []
    
    encabezados = [
        'ID ManyChat', 
        'RUT', 
        'WhatsApp', 
        'Email', 
        'Edad',
        'Comuna', 
        'PAP Alterado', 
        'Parejas sexuales'
    ]
    data.append(encabezados)

    for usuario in usuarios:
        edad = calcular_edad(usuario.fecha_nacimiento) if usuario.fecha_nacimiento else 'No registrada'
        
        fila = [
            usuario.id_manychat,
            f"{usuario.rut_usuario}-{usuario.dv_rut}",
            str(usuario.num_whatsapp),
            usuario.email or 'No disponible',
            str(edad),
            usuario.nombre_comuna,
            usuario.pap_alterado,
            usuario.parejas_sexuales
        ]
        data.append(fila)

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    
    tabla = Table(data)
    
    estilo = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#c6fffa')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
    ])
    
    for i in range(1, len(data)):
        if i % 2 == 0:
            estilo.add('BACKGROUND', (0, i), (-1, i), colors.whitesmoke)
    
    tabla.setStyle(estilo)
    
    ancho_columnas = [60, 80, 80, 120, 40, 100, 60, 80]
    for i, width in enumerate(ancho_columnas):
        estilo.add('COLWIDTH', (i, 0), (i, -1), width)
    
    elementos = []
    
    titulo = Paragraph("Listado Priorizado de Usuarios", styles['Title'])
    elementos.append(titulo)
    
    elementos.append(Paragraph("<br/><br/>", styles['Normal']))
    
    elementos.append(tabla)
    
    doc.build(elementos)
    
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="Listado_Priorizado.pdf"'
    
    return response
# -------------------------------------------------------------------- #
# ---------------------- Pregunta especialista ----------------------- #
# -------------------------------------------------------------------- #

@login_required
def preg_especialista(request):
    query = request.GET.get("q", "").strip()

    datos_query = UsuarioTextoPregunta.objects.select_related("id_manychat").order_by("-fecha_pregunta_texto")

    # Aplica el filtro si hay query
    if query:
        datos_query = filtrar_por_rut_o_manychat(datos_query, query)

    page_obj = paginacion_queryset1(request, datos_query)

    return render(request, "administracion/preg_especialista.html", {
        "page_obj": page_obj,
        "query": query
    })

def crear_excel_preg_especialista(request):
    session_key = request.GET.get('session_key')
    if session_key:
        session = SessionStore(session_key=session_key)
        if session.get('password_validated', False):
            
            wb = Workbook()
            ws_preg_esp = wb.active
            ws_preg_esp.title = "Preguntas especialistas"

            lista_preguntas = ['ID ManyChat', 'RUT', 'Pregunta', 'Fecha Pregunta']
            ws_preg_esp.append(lista_preguntas)

            preguntas = UsuarioTextoPregunta.objects.select_related('id_manychat').all()

            for pregunta in preguntas:
                fila = [
                    str(pregunta.id_manychat.id_manychat), 
                    f"{pregunta.id_manychat.rut_usuario}-{pregunta.id_manychat.dv_rut}",
                    pregunta.texto_pregunta,
                    pregunta.fecha_pregunta_texto.strftime('%d-%m-%Y %H:%M:%S') if pregunta.fecha_pregunta_texto else ''
                ]
                ws_preg_esp.append(fila)

            ajustar_ancho_columnas(ws_preg_esp)
            background_colors(ws_preg_esp)

            response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            response["Content-Disposition"] = 'attachment; filename="Preguntas_Especialista.xlsx"'

            wb.save(response)
            return response
    return HttpResponseForbidden("Acceso no autorizado")

  
@login_required
def crear_pdf_preg_especialista(request):
    session_key = request.GET.get('session_key')
    if session_key:
        session = SessionStore(session_key=session_key)
        if session.get('password_validated', False):
            session.flush()
            preguntas = UsuarioTextoPregunta.objects.select_related('id_manychat').all().order_by("-fecha_pregunta_texto")
            
            data = []
            
            encabezados = [
                'ID ManyChat', 
                'RUT', 
                'Pregunta', 
                'Fecha Pregunta'
            ]
            data.append(encabezados)

            for pregunta in preguntas:
                fila = [
                    pregunta.id_manychat.id_manychat,
                    f"{pregunta.id_manychat.rut_usuario}-{pregunta.id_manychat.dv_rut}",
                    pregunta.texto_pregunta,
                    pregunta.fecha_pregunta_texto.strftime('%d-%m-%Y %H:%M:%S') if pregunta.fecha_pregunta_texto else 'Sin fecha'
                ]
                data.append(fila)

            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=letter)
            styles = getSampleStyleSheet()
            
            tabla = Table(data)
            
            estilo = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#c6fffa')), 
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
            ])
            
            for i in range(1, len(data)):
                if i % 2 == 0:
                    estilo.add('BACKGROUND', (0, i), (-1, i), colors.whitesmoke)
            
            tabla.setStyle(estilo)
            
            ancho_columnas = [60, 80, 200, 100] 
            for i, width in enumerate(ancho_columnas):
                estilo.add('COLWIDTH', (i, 0), (i, -1), width)
            
            elementos = []
            
            titulo = Paragraph("Preguntas a Especialistas", styles['Title'])
            elementos.append(titulo)
            
            elementos.append(Paragraph("<br/><br/>", styles['Normal']))
            
            elementos.append(tabla)
            
            doc.build(elementos)
            
            buffer.seek(0)
            response = HttpResponse(buffer, content_type='application/pdf')
            response['Content-Disposition'] = 'attachment; filename="Preguntas_Especialistas.pdf"'
            
            return response
    return HttpResponseForbidden("Acceso no autorizado")

# ----------------------------------------------------------- #
# ---------------------- CRUD USUARIO ----------------------- #
# ----------------------------------------------------------- #
@login_required
def lista_usuarios(request):
    perfiles = PerfilUsuario.objects.select_related('user','usuario_sist').order_by('id_perfil')

    page_obj = paginacion_queryset1(request, perfiles)
    
    return render (request, 'administracion/lista_usuarios.html', {'page_obj': page_obj, 'perfiles': perfiles})

from django.contrib import messages

@login_required
def crear_usuario(request):
    if request.method == 'POST':
        form_user = UserForm(request.POST)
        form_perfil = PerfilUsuarioForm(request.POST)

        if form_user.is_valid() and form_perfil.is_valid():
            try:
                username = form_user.cleaned_data['username']
                if User.objects.filter(username=username).exists():
                    messages.error(request, 'El nombre de usuario ya existe.')
                    return render(request, 'administracion/form_usuario.html', {
                        'form_user': form_user,
                        'form_perfil': form_perfil,
                        'creando': True
                    })

                tipo = form_perfil.cleaned_data.get('tipo_usuario')

                if tipo == 'paciente':
                    email = form_user.cleaned_data.get('email')
                    telefono = form_perfil.cleaned_data.get('telefono')
                    rut = form_perfil.cleaned_data.get('rut_usuario')
                    dv = form_perfil.cleaned_data.get('dv_rut')
                    usuario_sist = form_perfil.cleaned_data.get('usuario_sist')

                    errores = []

                    if not Usuario.objects.filter(id_manychat=usuario_sist.id_manychat).exists():
                        errores.append("ID ManyChat no coincide")
                    else:
                        usuario = Usuario.objects.get(id_manychat=usuario_sist.id_manychat)

                        if usuario.rut_usuario != rut:
                            errores.append("RUT no coincide")
                        if usuario.dv_rut.upper() != dv.upper():
                            errores.append("Dígito verificador no coincide")
                        if usuario.num_whatsapp != int(telefono):
                            errores.append(f"Número de teléfono no coincide)")
                        if usuario.email.strip().lower() != email.strip().lower():
                            errores.append("Correo electrónico no coincide")

                    if errores:
                        print("VALIDACIÓN FALLIDA:\n", "\n".join(errores))
                        messages.error(request, " - ".join(errores))
                        return render(request, 'administracion/form_usuario.html', {
                            'form_user': form_user,
                            'form_perfil': form_perfil,
                            'creando': True
                        })

                user = form_user.save()
                perfil = form_perfil.save(commit=False)
                perfil.user = user
                perfil.save()

                messages.success(request, 'Usuario creado correctamente.')
                return redirect('lista_usuarios')

            except Exception as e:
                messages.error(request, f'Error al crear usuario: {str(e)}')
        else:
            messages.error(request, 'Por favor corrija los errores en el formulario.')

    else:
        form_user = UserForm()
        form_perfil = PerfilUsuarioForm()

    return render(request, 'administracion/form_usuario.html', {
        'form_user': form_user,
        'form_perfil': form_perfil,
        'creando': True
    })

@login_required
def editar_usuario(request, perfil_id):
    perfil = get_object_or_404(PerfilUsuario, id_perfil=perfil_id)
    user = perfil.user

    if request.method == 'POST':
        form_user = UserForm(request.POST, instance=user)
        form_perfil = PerfilUsuarioForm(request.POST, instance=perfil)

        if form_user.is_valid() and form_perfil.is_valid():
            try:
                email = form_user.cleaned_data['email']
                if User.objects.filter(email=email).exclude(pk=user.pk).exists():
                    messages.error(request, 'El correo electrónico ya está en uso.')
                    return render(request, 'administracion/form_usuario.html', {
                        'form_user': form_user,
                        'form_perfil': form_perfil
                    })

                form_user.save()
                form_perfil.save()
                
                messages.success(request, 'Usuario actualizado correctamente.')
                return redirect('lista_usuarios')
            except Exception as e:
                messages.error(request, f'Error al actualizar usuario: {str(e)}')
        else:
            messages.error(request, 'Por favor corrija los errores en el formulario.')
    else:
        form_user = UserForm(instance=user)
        form_perfil = PerfilUsuarioForm(instance=perfil)

    return render(request, 'administracion/form_usuario.html', {
        'form_user': form_user,
        'form_perfil': form_perfil,
        'creando': False
    })

@login_required
def eliminar_usuario(request, perfil_id):
    perfil = get_object_or_404(PerfilUsuario, id_perfil=perfil_id)
    if request.method == 'POST':
        perfil.user.delete()
        perfil.delete()
        messages.success(request, 'Usuario eliminado.')
        return redirect('lista_usuarios')
    return render(request, 'administracion/confirmar_eliminacion.html', {'perfil': perfil})
  
# ------------------------------------------------------------------------- #
# ---------------------- Historial de agendamientos ----------------------- #
# ------------------------------------------------------------------------- #

@login_required
def generar_json_por_cesfam(request, cesfam_id):
    try:
        cesfam = Cesfam.objects.get(id_cesfam=cesfam_id)
        agendas = Agenda.objects.filter(id_cesfam=cesfam_id).select_related(
            'id_manychat', 'id_procedimiento', 'id_manychat__perfilusuario__user'
        )
        
        if not agendas.exists():
            return JsonResponse({"error": "No hay horas agendadas para este CESFAM"}, status=404)
        
        datos_agendas = []
        for agenda in agendas:
            nombre_paciente = "Paciente No Registrado"
            rut_paciente = f"{agenda.id_manychat.rut_usuario}-{agenda.id_manychat.dv_rut}" if agenda.id_manychat.rut_usuario else "Sin RUT"
            
            try:
                if hasattr(agenda.id_manychat, 'perfilusuario') and agenda.id_manychat.perfilusuario.user:
                    user = agenda.id_manychat.perfilusuario.user
                    nombre_paciente = user.get_full_name() or f"{user.first_name} {user.last_name}".strip()
            except Exception:
                pass
            
            if nombre_paciente == "Paciente No Registrado":
                nombre_paciente = getattr(agenda.id_manychat, 'nombre_whatsapp', "Paciente ManyChat")
            
            datos_agendas.append({
                "rut": rut_paciente,
                "nombre": nombre_paciente,
                "fecha": agenda.fecha_atencion.strftime("%d-%m-%Y"),
                "hora": agenda.hora_atencion.strftime("%H:%M"),
                "procedimiento": agenda.id_procedimiento.nombre_procedimiento,
                "contacto": str(agenda.id_manychat.num_whatsapp),
                "manychat_id": agenda.id_manychat.id_manychat,
                "tipo_usuario": "registrado" if hasattr(agenda.id_manychat, 'perfilusuario') else "no_registrado"
            })
        
        response_data = {
            "metadata": {
                "sistema": "APTRedTami",
                "version": "1.0",
                "fecha_generacion": datetime.now().strftime("%d-%m-%Y %H:%M:%S"),
                "total_horas": len(datos_agendas),
                "horas_registradas": len([a for a in datos_agendas if a['tipo_usuario'] == "registrado"])
            },
            "cesfam": {
                "id": cesfam.id_cesfam,
                "nombre": cesfam.nombre_cesfam,
                "comuna": cesfam.cod_comuna.nombre_comuna
            },
            "horas_agendadas": datos_agendas
        }

        nombre_archivo = f"horas_cesfam_{cesfam.id_cesfam}_{datetime.now().strftime('%Y%m%d')}.json"
        LogDescargaJSON.objects.create(
            usuario=request.user,
            cesfam=cesfam,
            cantidad_horas=len(datos_agendas),
            nombre_archivo=nombre_archivo
        )

        response = HttpResponse(
            json.dumps(response_data, indent=2, ensure_ascii=False),
            content_type='application/json; charset=utf-8'
        )
        response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
        return response

    except Cesfam.DoesNotExist:
        return JsonResponse({"error": "CESFAM no encontrado"}, status=404)

@login_required
def lista_descargas(request):
    cesfams = Cesfam.objects.annotate(
        num_horas=Count('agenda')
    )
    return render(request, 'administracion/json_cesfam.html', {'cesfams': cesfams})

@login_required
def historial_descargas_json(request):
    descargas_queryset = LogDescargaJSON.objects.select_related('usuario', 'cesfam').order_by('-fecha_descarga')
    page_obj = paginacion_queryset1(request, descargas_queryset, items_por_pagina=20)  

    return render(request, 'administracion/historial_descargas.html', {'page_obj': page_obj})

def es_administrador(user):
    return hasattr(user, 'perfilusuario') and user.perfilusuario.tipo_usuario == 'administrador'

@login_required(login_url='/login/')
@user_passes_test(es_administrador, login_url='/login/')
def exportar_historial_excel(request):
    search_query = request.GET.get('search', '')

    agendamientos = Agenda.objects.select_related(
        'id_manychat', 
        'id_cesfam', 
        'id_procedimiento'
    ).filter(
        Q(id_manychat__rut_usuario__icontains=search_query) |
        Q(id_manychat__email__icontains=search_query) |
        Q(id_manychat__id_manychat__icontains=search_query)
    ).order_by('-fecha_atencion')

    wb = Workbook()  
    ws = wb.active
    ws.title = "Historial de Agendamientos"
    
    # Encabezados
    columnas = [
        'RUT', 
        'CESFAM', 
        'Procedimiento', 
        'Requisitos',
        'Fecha', 
        'Hora', 
        'Email'
    ]
    ws.append(columnas)

    # Datos
    for agendamiento in agendamientos:
        ws.append([
            f"{agendamiento.id_manychat.rut_usuario}-{agendamiento.id_manychat.dv_rut}",
            agendamiento.id_cesfam.nombre_cesfam,
            agendamiento.id_procedimiento.nombre_procedimiento,
            agendamiento.requisito_examen,
            agendamiento.fecha_atencion.strftime('%d-%m-%Y'),
            agendamiento.hora_atencion.strftime('%H:%M'),
            agendamiento.id_manychat.email or 'No registrado'
        ])

    # Preparar la respuesta
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=historial_agendamientos.xlsx'
    return response

@login_required(login_url='/login/')
@user_passes_test(es_administrador, login_url='/login/')
def historial_agendamientos(request):
    search_query = request.GET.get('search', '').strip()

    agendamientos = Agenda.objects.select_related('id_manychat', 'id_cesfam', 'id_procedimiento')

    # Anotamos el rut completo concatenando rut_usuario + '-' + dv_rut
    agendamientos = agendamientos.annotate(
        rut_completo=Concat(
            F('id_manychat__rut_usuario'),
            Value('-'),
            F('id_manychat__dv_rut'),
            output_field=CharField()
        )
    )

    if search_query:
        # Limpiamos la búsqueda para eliminar puntos, guiones o espacios (opcional)
        cleaned_query = search_query.replace(".", "").replace("-", "").replace(" ", "").lower()

        # Filtramos por rut_completo con y sin formato, email o nombre cesfam
        agendamientos = agendamientos.filter(
            Q(rut_completo__icontains=search_query) |  # con guion
            Q(rut_completo__icontains=cleaned_query) |  # sin guion y espacios
            Q(id_manychat__email__icontains=search_query) |
            Q(id_cesfam__nombre_cesfam__icontains=search_query)
        )

    agendamientos = agendamientos.order_by('-fecha_atencion')

    page_obj = paginacion_queryset1(request, agendamientos, items_por_pagina=10)

    context = {
        'page_obj': page_obj,
        'search_query': search_query,
    }
    return render(request, 'administracion/historial_agendamientos.html', context)